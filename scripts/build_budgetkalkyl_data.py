#!/usr/bin/env python3
"""Exportera publik budgetdata från Excel-mastern utan att ändra arbetsboken."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Fel: openpyxl saknas. Installera openpyxl i Python-miljön och kör igen.") from exc

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data" / "budgetkalkylator" / "AVAB-budgetkalkyl-master.xlsx"
OUTPUT = ROOT / "src" / "data" / "budgetkalkylator" / "priser.generated.json"
SHEET = "Webbexport"
REQUIRED = ("Område", "Grupp", "Nyckel", "Paket-ID", "Visningsnamn", "Typ", "Pris", "Timmar", "Offert?", "Notering")
AREA_META = {
    "Sporthall": ("sporthall", "Sporthall och arena", False),
    "Simhall": ("simhall", "Simhall", True),
    "Konferens": ("konferens", "Konferens", True),
    "Restaurang": ("restaurang", "Restaurang", True),
    "Gym": ("gym", "Gym", True),
}


def yes(value: object) -> bool:
    return str(value or "").strip().casefold() in {"ja", "true", "1", "yes"}


def number(value: object, field: str, row: int) -> float | int:
    if value in (None, ""):
        return 0
    try:
        parsed = float(str(value).replace(",", "."))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Rad {row}: {field} måste vara ett tal, fick {value!r}.") from exc
    return int(parsed) if parsed.is_integer() else parsed


def main() -> None:
    if not WORKBOOK.is_file():
        raise FileNotFoundError(f"Excel-master saknas: {WORKBOOK}")
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    missing_sheets = [name for name in ("Start", "Inställningar", *AREA_META, SHEET) if name not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError("Arbetsboken saknar flik(ar): " + ", ".join(missing_sheets))
    sheet = workbook[SHEET]
    header_row = next((row for row in range(1, min(sheet.max_row, 30) + 1) if sheet.cell(row, 1).value == "Område"), None)
    if not header_row:
        raise ValueError("Webbexport saknar rubrikraden som börjar med 'Område'.")
    headers = {sheet.cell(header_row, col).value: col for col in range(1, sheet.max_column + 1)}
    missing = [name for name in REQUIRED if name not in headers]
    if missing:
        raise ValueError("Webbexport saknar obligatoriska kolumner: " + ", ".join(missing))

    areas = {slug: {"label": label, "isTestValue": test, "groups": {}} for slug, label, test in AREA_META.values()}
    seen_ids: set[tuple[str, str]] = set()
    exported = 0
    for row in range(header_row + 1, sheet.max_row + 1):
        area_name = sheet.cell(row, headers["Område"]).value
        if area_name in (None, ""):
            continue
        if area_name not in AREA_META:
            raise ValueError(f"Rad {row}: okänt område {area_name!r}.")
        values = {name: sheet.cell(row, col).value for name, col in headers.items() if name in REQUIRED}
        group = str(values["Grupp"] or "").strip()
        key = str(values["Nyckel"] or "").strip()
        package_id = str(values["Paket-ID"] or "").strip()
        if not group or not key or not package_id:
            raise ValueError(f"Rad {row}: Grupp, Nyckel och Paket-ID måste vara ifyllda.")
        if package_id != f"{group}:{key}":
            raise ValueError(f"Rad {row}: Paket-ID {package_id!r} matchar inte {group}:{key}.")
        slug = AREA_META[area_name][0]
        identity = (slug, package_id)
        if identity in seen_ids:
            raise ValueError(f"Rad {row}: duplicerat Paket-ID {package_id!r} i {area_name}.")
        seen_ids.add(identity)
        price = number(values["Pris"], "Pris", row)
        hours = number(values["Timmar"], "Timmar", row)
        note = str(values["Notering"] or "").strip()
        item = {
            "label": str(values["Visningsnamn"] or "").strip(),
            "type": str(values["Typ"] or "").strip(),
            "price": price,
            "hours": hours,
            "quote": yes(values["Offert?"]),
            "note": note,
        }
        if group == "standAccess":
            extra = re.search(r"Extra\s+([0-9]+(?:[.,][0-9]+)?)\s*kr", note, re.IGNORECASE)
            item = {"multiplier": price, "extra": number(extra.group(1), "Extra", row) if extra else 0, "hourMultiplier": hours, "note": note}
        elif group == "ledPitch":
            start = re.search(r"Startpris\s+([0-9]+(?:[.,][0-9]+)?)\s*kr", note, re.IGNORECASE)
            item = {"label": str(values["Visningsnamn"] or "").strip(), "sqm": price, "start": number(start.group(1), "Startpris", row) if start else 0}
            if yes(values["Offert?"]):
                item["quote"] = True
        areas[slug]["groups"].setdefault(group, {})[key] = item
        exported += 1

    if not exported:
        raise ValueError("Webbexport innehåller inga exporterbara rader.")
    payload = {
        "meta": {"currency": "SEK", "vat": "exclusive", "warrantyRate": 0.10},
        "areas": areas,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Klart: exporterade {exported} rader till {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Fel vid budgetexport: {exc}", file=sys.stderr)
        raise SystemExit(1)
